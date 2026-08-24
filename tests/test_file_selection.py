from __future__ import annotations

from agent.discovery.file_family import FileFamily, freeze_file_selection
from agent.discovery.file_export import write_file_selection_exports
from agent.discovery.file_judgment import (
    FileJudgmentInput,
    check_file_evidence,
    is_file_selection_ready,
    stable_file_id,
)
from agent.discovery.file_review import (
    file_evidence_values,
    generate_selected_file_reasons,
    review_files_in_batches,
)
from agent.discovery.models import DatasetManifest, DatasetRequest, DiscoveredFile, DiscoveredProject


def _judgment(
    file: DiscoveredFile,
    *,
    role: str = "primary_input",
    companions: list[str] | None = None,
) -> FileJudgmentInput:
    return FileJudgmentInput(
        file_id=file.file_id,
        project_accession=file.project_accession,
        file_name=file.file_name,
        selection_role=role,
        review_status="reviewed",
        decision="include",
        grade=3,
        hard_gate="pass",
        confidence=0.91,
        evidence_refs=[f"{file.file_id}:metadata"],
        companion_file_ids=companions or [],
        reason_outline=["目标样本原始质谱文件", "SDRF 映射完整"],
        reason_text=f"选择 {file.file_name}，因为文件级证据与任务要求一致。",
        reason_status="ready",
        model_id="test-model",
    )


def test_file_judgment_evidence_and_companion_selection_are_file_scoped() -> None:
    raw = DiscoveredFile(
        repository="pride",
        project_accession="PXD000001",
        file_accession_or_path="RAW-1",
        file_name="sample-1.raw",
        file_type=".raw",
    )
    sdrf = DiscoveredFile(
        repository="pride",
        project_accession="PXD000001",
        file_accession_or_path="SDRF-1",
        file_name="experiment.sdrf.tsv",
        file_type=".tsv",
        file_role="metadata",
        selection_role="required_companion",
    )
    assert raw.file_id == stable_file_id("pride", "PXD000001", "RAW-1")
    assert raw.file_id != sdrf.file_id

    raw_judgment = _judgment(raw, companions=[sdrf.file_id])
    sdrf_judgment = _judgment(sdrf, role="required_companion")
    evidence_check = check_file_evidence(
        raw_judgment,
        [f"{raw.file_id}:metadata"],
    )
    assert is_file_selection_ready(raw_judgment, evidence_check)

    family = FileFamily(
        family_id="family-1",
        project_accession="PXD000001",
        primary_file_ids=[raw.file_id],
        companion_file_ids=[sdrf.file_id],
    )
    selected, blockers = freeze_file_selection(
        [raw.file_id],
        {raw.file_id: raw_judgment, sdrf.file_id: sdrf_judgment},
        [family],
    )
    assert selected == [raw.file_id, sdrf.file_id]
    assert blockers == []


def test_llm_review_batches_files_then_writes_long_reasons_only_for_selected() -> None:
    files = [
        DiscoveredFile(
            repository="pride",
            project_accession="PXD000001",
            file_accession_or_path=f"RAW-{index}",
            file_name=f"sample-{index}.raw",
            file_type=".raw",
        )
        for index in range(2)
    ]

    class FakeLLM:
        model = "fake-scientific-model"

        def complete_json(self, *, system_prompt: str, user_prompt: str):
            if "reason paragraph" in user_prompt:
                return {
                    "reasons": [
                        {
                            "file_id": file.file_id,
                            "reason_text": f"选择 {file.file_name}，因为文件证据支持该任务。",
                        }
                        for file in files
                    ]
                }
            return {
                "judgments": [
                    {
                        "file_id": file.file_id,
                        "selection_role": "primary_input",
                        "review_status": "reviewed",
                        "decision": "include",
                        "grade": 3,
                        "hard_gate": "pass",
                        "confidence": 0.9,
                        "evidence_refs": [next(iter(file_evidence_values(file)))],
                        "reason_outline": ["文件证据符合任务"],
                        "reason_status": "pending",
                    }
                    for file in files
                ]
            }

    request = DatasetRequest(query_terms=["human"])
    judgments = review_files_in_batches(
        FakeLLM(),
        request=request,
        projects=[
            DiscoveredProject(
                project_accession="PXD000001",
                project_title="Human proteomics",
            )
        ],
        files=files,
    )
    completed = generate_selected_file_reasons(
        FakeLLM(),
        request=request,
        files=files,
        judgments=judgments,
    )

    assert set(completed) == {file.file_id for file in files}
    assert all(item.reason_status == "ready" for item in completed.values())
    assert all(item.reason_text and item.file_name in item.reason_text for item in completed.values())


def test_exports_keep_excel_small_and_preserve_excluded_reason(tmp_path) -> None:
    from openpyxl import load_workbook
    import pyarrow.parquet as parquet

    selected = DiscoveredFile(
        repository="pride",
        project_accession="PXD000010",
        file_accession_or_path="selected.raw",
        file_name="selected.raw",
        file_type=".raw",
        decision="include",
        reason_text="Selected file reason.",
    )
    excluded = DiscoveredFile(
        repository="pride",
        project_accession="PXD000010",
        file_accession_or_path="excluded.raw",
        file_name="excluded.raw",
        file_type=".raw",
        decision="exclude",
        reason_text="Excluded file reason.",
    )
    manifest = DatasetManifest(
        request=DatasetRequest(query_terms=["human"]),
        files=[selected, excluded],
        summary={
            "all_file_judgments": {
                selected.file_id: _judgment(selected).model_dump(mode="json"),
                excluded.file_id: FileJudgmentInput(
                    file_id=excluded.file_id,
                    project_accession=excluded.project_accession,
                    file_name=excluded.file_name,
                    review_status="reviewed",
                    decision="exclude",
                    grade=0,
                    hard_gate="fail",
                    confidence=0.95,
                    reason_text="Excluded file reason.",
                    reason_status="ready",
                ).model_dump(mode="json"),
            }
        },
    )

    paths = write_file_selection_exports(manifest, tmp_path)
    workbook = load_workbook(paths["selected_files_xlsx"], read_only=True)
    selected_names = [row[3] for row in workbook["Selected Files"].iter_rows(min_row=2, values_only=True)]
    judgment_rows = parquet.read_table(paths["file_judgments_parquet"]).to_pylist()

    assert selected_names == ["selected.raw"]
    reasons = {row["reason_text"] for row in judgment_rows}
    assert len(reasons) == 2
    assert "Excluded file reason." in reasons
