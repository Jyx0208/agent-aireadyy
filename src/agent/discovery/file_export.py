from __future__ import annotations

import json
from collections.abc import Iterable
from pathlib import Path
from typing import Any

import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill

from agent.discovery.models import DatasetManifest, DiscoveredFile


FILE_EXPORT_COLUMNS = (
    "file_id",
    "repository",
    "project_accession",
    "file_name",
    "file_type",
    "file_role",
    "selection_role",
    "family_id",
    "decision",
    "grade",
    "confidence",
    "reason_text",
    "companion_file_ids",
    "download_url",
)


def _judgments(manifest: DatasetManifest) -> dict[str, dict[str, Any]]:
    raw = (
        manifest.summary.get("all_file_judgments")
        or manifest.summary.get("file_judgments")
        or {}
    )
    return raw if isinstance(raw, dict) else {}


def _row(file: DiscoveredFile, judgments: dict[str, dict[str, Any]]) -> dict[str, Any]:
    judgment = judgments.get(file.file_id) or {}
    return {
        "file_id": file.file_id,
        "repository": file.repository,
        "project_accession": file.project_accession,
        "file_name": file.file_name,
        "file_type": file.file_type,
        "file_role": file.file_role,
        "selection_role": judgment.get("selection_role", file.selection_role),
        "family_id": judgment.get("family_id", file.family_id),
        "decision": judgment.get("decision", file.decision),
        "grade": judgment.get("grade"),
        "confidence": judgment.get("confidence"),
        "reason_text": judgment.get("reason_text", file.reason_text),
        "companion_file_ids": json.dumps(
            judgment.get("companion_file_ids") or file.companion_file_ids,
            ensure_ascii=False,
        ),
        "download_url": file.download_url,
    }


def _write_parquet(path: Path, rows: Iterable[dict[str, Any]], batch_size: int = 2_000) -> None:
    writer: pq.ParquetWriter | None = None
    batch: list[dict[str, Any]] = []
    for row in rows:
        batch.append(row)
        if len(batch) < batch_size:
            continue
        table = pa.Table.from_pylist(batch)
        writer = writer or pq.ParquetWriter(path, table.schema, compression="zstd")
        writer.write_table(table)
        batch.clear()
    if batch:
        table = pa.Table.from_pylist(batch)
        writer = writer or pq.ParquetWriter(path, table.schema, compression="zstd")
        writer.write_table(table)
    if writer is None:
        pq.write_table(pa.table({"file_id": pa.array([], type=pa.string())}), path)
    else:
        writer.close()


def _header_cells(sheet: Any) -> list[WriteOnlyCell]:
    cells: list[WriteOnlyCell] = []
    for value in FILE_EXPORT_COLUMNS:
        cell = WriteOnlyCell(sheet, value=value)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="161616")
        cells.append(cell)
    return cells


def _append_file_sheet(sheet: Any, rows: Iterable[dict[str, Any]]) -> None:
    sheet.append(_header_cells(sheet))
    for row in rows:
        sheet.append([row.get(column) for column in FILE_EXPORT_COLUMNS])


def write_file_selection_exports(
    manifest: DatasetManifest,
    output_dir: str | Path,
) -> dict[str, Path]:
    """Write a small human workbook and scalable machine-readable tables."""

    output = Path(output_dir)
    output.mkdir(parents=True, exist_ok=True)
    judgments = _judgments(manifest)
    rows = [_row(file, judgments) for file in manifest.files]
    selected_rows = [row for row in rows if row["decision"] in {None, "include"}]
    primary_rows = [row for row in selected_rows if row["selection_role"] == "primary_input"]
    companion_rows = [row for row in selected_rows if row["selection_role"] != "primary_input"]

    workbook_path = output / "selected_files.xlsx"
    workbook = Workbook(write_only=True)
    readme = workbook.create_sheet("README")
    readme.append(["用途", "最终入选文件清单；完整评审数据请使用 Parquet 或 JSONL。"])
    readme.append(["选择单位", "文件，而不是整个项目。"])
    readme.append(["SDRF", "与入选原始文件明确匹配时，列入 Companions。"])
    selected_sheet = workbook.create_sheet("Selected Files")
    companion_sheet = workbook.create_sheet("Companions")
    _append_file_sheet(selected_sheet, primary_rows)
    _append_file_sheet(companion_sheet, companion_rows)
    workbook.save(workbook_path)

    files_path = output / "files.parquet"
    judgments_path = output / "file_judgments.parquet"
    evidence_path = output / "file_evidence.parquet"
    _write_parquet(files_path, rows)
    _write_parquet(
        judgments_path,
        (
            {
                **judgment,
                "evidence_refs": json.dumps(judgment.get("evidence_refs") or [], ensure_ascii=False),
                "limitations": json.dumps(judgment.get("limitations") or [], ensure_ascii=False),
                "missing_information": json.dumps(judgment.get("missing_information") or [], ensure_ascii=False),
                "companion_file_ids": json.dumps(judgment.get("companion_file_ids") or [], ensure_ascii=False),
                "reason_outline": json.dumps(judgment.get("reason_outline") or [], ensure_ascii=False),
            }
            for judgment in judgments.values()
        ),
    )
    _write_parquet(
        evidence_path,
        (
            {
                "file_id": file.file_id,
                "evidence_index": index,
                **evidence.model_dump(mode="json"),
            }
            for file in manifest.files
            for index, evidence in enumerate(file.evidence)
        ),
    )
    return {
        "selected_files_xlsx": workbook_path,
        "files_parquet": files_path,
        "file_judgments_parquet": judgments_path,
        "file_evidence_parquet": evidence_path,
    }
